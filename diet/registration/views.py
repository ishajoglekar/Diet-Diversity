import io
import ast
from typing import NewType
from django.contrib.auth.models import Group
from django.db.models.expressions import F

import openpyxl
import string
import random
import xlsxwriter
from django.shortcuts import render,redirect
from django.contrib import messages
from django.http import HttpResponse
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import login,authenticate
from datetime import datetime,date
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required,user_passes_test
from django.core.exceptions import PermissionDenied
from django.urls import reverse
from .decorators import *

from registration.models import *
from .forms import *
from shared.encryption import EncryptionHelper

#user to check if a user belongs to a group
def is_member(user,grp):
    grp = Group.objects.get(pk=grp)
    return user.groups.filter(name=grp).exists()

def is_student(user):
    return user.groups.filter(name='Students').exists()

def is_parent(user):
    return user.groups.filter(name='Parents').exists()

def is_teacher(user):
    return user.groups.filter(name='Teachers').exists()

def home(request):
    return render(request, "registration_form/home.html")

def consent(request):
    if request.method == "GET":
        form = ConsentForm()
        return render(request,'registration_form/consent.html',{'form':form})

def parents_info(request):
    if request.method == "GET":
        if(request.session.get('data')):
            form = ParentsInfoForm(request.session.get('data'))
            user_creation_form = UserCreationForm(request.session.get('data'))
        else:
            form = ParentsInfoForm()
            user_creation_form = UserCreationForm() 
        return render(request,'registration_form/parents_info.html',{'form':form,'user_creation_form':user_creation_form})
    else:
        form = ParentsInfoForm(request.POST)
        user_creation_form =  UserCreationForm(request.POST)

        if form.is_valid() and user_creation_form.is_valid():       
            print(request.POST['state'].strip())     
            print(request.POST['city'].strip())     
            request.session['data'] = request.POST
            return redirect('/students_info')
        else:                        
            return render(request,'registration_form/parents_info.html',{'form':form,'user_creation_form':user_creation_form})
    

def students_info(request):
    if request.method == "GET":
        form = StudentsInfoForm()
        user_creation_form = UserCreationForm() 
        return render(request,'registration_form/students_info.html',{'form':form,'user_creation_form':user_creation_form})
    else:        
        previousPOST = request.session['data']
        form = StudentsInfoForm(request.POST)
        studentuserform =  UserCreationForm(request.POST)
        parentform = ParentsInfoForm(previousPOST)
        parentuserform =  UserCreationForm(previousPOST)
        if form.is_valid() and studentuserform.is_valid():                        
            encryptionHelper = EncryptionHelper()
            parentUser = parentuserform.save(commit=False)
            parentUser.save()
            parent_group = Group.objects.get(name='Parents')
            parentUser.groups.add(parent_group)
            parentUser.save()

            parent = parentform.save(commit=False)    
            parent.user = parentUser
            parent.first_password = ''
            parent.password_changed = True                       
            parent.name = encryptionHelper.encrypt(previousPOST['name'])
            parent.email = encryptionHelper.encrypt(previousPOST['email'])
            parent.state = State.objects.get(state__icontains=previousPOST['state'].strip())
            parent.city = City.objects.get(city__icontains=previousPOST['city'].strip())
            parent.save()

            studentuser = studentuserform.save(commit=False)
            studentuser.save()
            student_group = Group.objects.get(name='Students')
            studentuser.groups.add(student_group)
            studentuser.save()

            student = form.save(commit=False)
            student.user = studentuser            
            student.first_password = ''
            student.password_changed = True
            student.name = encryptionHelper.encrypt(request.POST['name'])
            student.parent = parent
            student.save()


            user = authenticate(request, username=previousPOST['username'], password=previousPOST['password1'])
            if user is not None:
                login(request, user)

            if 'data' in request.session:
                del request.session['data']
            return redirect('/parent_dashboard')
        else:                        
            return render(request,'registration_form/students_info.html',{'form':form,'user_creation_form':studentuserform})


def loginU(request):
    if request.method == "GET":
        form = CustomAuthenticationForm()
        return render(request,'registration_form/login.html',{'form':form})
    else:
        username = request.POST['username']
        password = request.POST['password']
        grp = request.POST['groups']
        user = authenticate(request, username=username, password=password)
        grp_name = Group.objects.get(pk=grp).name
        form = CustomAuthenticationForm(request.POST)
        if user is not None:
            if is_member(user,grp):
                login(request,user)                
                if grp_name == 'Parents':
                    return redirect('/parent_dashboard')
                elif grp_name == 'Students':
                    return redirect('/student_dashboard')
                elif grp_name == 'Teachers':
                    return redirect('/teacher_dashboard')
            else:
                messages.error(request, 'User does not belong to selected group')
                return render(request,'registration_form/login.html',{'form':form})
        else:
            messages.error(request, 'Invalid credentials')
            return render(request,'registration_form/login.html',{'form':form})

@login_required(login_url='/login')
@user_passes_test(is_parent,login_url='/forbidden')
def dashboard(request):
    if request.method == "GET":
        students = ParentsInfo.objects.filter(user= request.user).first().studentsinfo_set.all()
        helper = EncryptionHelper()
        for student in students:    
            print(student.name)
            student.name = helper.decrypt(student.name)
        return render(request,'registration_form/dashboard.html',{'students':students})


@login_required(login_url='/login')
def logoutU(request):
    logout(request)
    return redirect('/login')

@login_required(login_url='/login')             
@user_passes_test(is_parent,login_url='/forbidden')
def addStudentForm(request):
    if request.method == "GET":
        form = StudentsInfoForm()
        user_creation_form = UserCreationForm() 
        return render(request,'registration_form/add_student.html',{'form':form,'user_creation_form':user_creation_form})
    else:
        form = StudentsInfoForm(request.POST)
        studentuserform =  UserCreationForm(request.POST)
        if form.is_valid() and studentuserform.is_valid():                        
            encryptionHelper = EncryptionHelper()
            studentuser = studentuserform.save(commit=False)
            studentuser.save()   
            student_group = Group.objects.get(name='Students')
            studentuser.groups.add(student_group)
            studentuser.save()

            student = form.save(commit=False)
            student.user = studentuser
            student.first_password = ''
            student.password_changed = True
            student.name = encryptionHelper.encrypt(request.POST['name']) 
            print(student.name)           
            student.parent = ParentsInfo.objects.filter(user= request.user).first()
            student.save()
            return redirect('/parent_dashboard')
        else:                        
            return render(request,'registration_form/add_student.html',{'form':form,'user_creation_form':studentuserform})


@login_required(login_url='/login')
@user_passes_test(is_teacher,login_url='/forbidden')
def bulkRegister(request):
    if(request.method=="GET"):
        return render(request,'registration/bulkregistration.html')
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        parentSheet = wb["Parents Data"]
        studentSheet = wb["Students Data"]

        encryptionHelper = EncryptionHelper()

        parent_data = list()
        # iterating over the rows and
        # getting value from each cell in row        
        for row in parentSheet.iter_rows():
            row_data = list()
            for cell in row:                                
                if cell.row == 1 :                
                    continue
                if cell.column_letter == 'A':                    
                    row_data.append(encryptionHelper.encrypt(str(cell.value)))
                elif cell.column_letter == 'B':
                    row_data.append(encryptionHelper.encrypt(str(cell.value)))
                    row_data.append(str(cell.value.lower().replace(" ",""))+str(random.randint(11,99)))
                elif cell.column_letter == 'C':
                    row_data.append(str(cell.value))
                elif cell.column_letter == 'D':
                    row_data.append(str(cell.value))
                elif cell.column_letter == 'E':
                    row_data.append(str(cell.value))
                elif cell.column_letter == 'F':
                    row_data.append(int(cell.value))
                elif cell.column_letter == 'G':
                    row_data.append(int(cell.value))
                elif cell.column_letter == 'H':
                    row_data.append(int(cell.value))
                elif cell.column_letter == 'I':
                    row_data.append(str(cell.value))
                elif cell.column_letter == 'J':
                    row_data.append(str(cell.value))
                elif cell.column_letter == 'K':
                    row_data.append(str(cell.value))
                elif cell.column_letter == 'L':
                    row_data.append(str(cell.value))
                elif cell.column_letter == 'M':
                    row_data.append(str(cell.value))
                elif cell.column_letter == 'N':
                    row_data.append(str(cell.value))
            parent_data.append(row_data)

        student_data = list()
        # iterating over the rows and
        # getting value from each cell in row                      
        for row in studentSheet.iter_rows():
            row_data = list()
            for cell in row:                                
                if cell.row == 1 :                
                    continue
                if cell.column_letter == 'A':                    
                    row_data.append(encryptionHelper.encrypt(str(cell.value)))
                    row_data.append(str(cell.value.lower().replace(" ",""))+str(random.randint(11,99)))
                elif cell.column_letter == 'B':
                    row_data.append(str(cell.value))
                elif cell.column_letter == 'C':
                    row_data.append(int(cell.value))
                elif cell.column_letter == 'D':
                    row_data.append(str(cell.value))
                elif cell.column_letter == 'E':
                    row_data.append(str(cell.value))
                elif cell.column_letter == 'F':
                    row_data.append(str(cell.value))
                elif cell.column_letter == 'G':
                    row_data.append(str(cell.value))
            student_data.append(row_data)

        
        for index,row in enumerate(parent_data):
            if index == 0:                
                continue  
            #creating parent user
            skipparent = False
            parentData = ParentsInfo.objects.all()
            for parent in parentData:                
                if encryptionHelper.decrypt(parent.email) == encryptionHelper.decrypt(row[0]):
                    skipparent = True
                    break
               
            if not skipparent:              
                password = ''.join(random.choices(string.ascii_lowercase + string.digits, k = 8))                      
                parentUser = User(username=row[2])
                parentUser.set_password(password)
                parentUser.save()
                parent_group = Group.objects.get(name='Parents')
                parentUser.groups.add(parent_group)
                parentUser.save()

                #getting data from db for foreign keys
                city = City.objects.filter(city__icontains=row[9]).first()
                state = State.objects.filter(state__icontains=row[10]).first()
                education = Education.objects.filter(education__icontains=row[11]).first()
                occupation = Occupation.objects.filter(occupation__icontains=row[12]).first()
                religion = ReligiousBelief.objects.filter(religion__icontains=row[13]).first()
                familyType = FamilyType.objects.filter(family__icontains=row[14]).first()  
                #creating parent
                parent = ParentsInfo(email=row[0],name=row[1],gender=row[3],age=row[4],address=row[5],pincode=row[6],no_of_family_members=row[7],children_count=row[8],city=city,state=state,edu=education,occupation=occupation,religion=religion,type_of_family=familyType,first_password=password)
                parent.user = parentUser                
                parent.save()
        
        
        for index,row in enumerate(student_data):
            if index == 0:
                continue  
            #creating student user
            skipstudent = StudentsInfo.objects.filter(rollno = row[3]).first()                
            if not skipstudent:               
                password = ''.join(random.choices(string.ascii_lowercase + string.digits, k = 8))
                studentUser = User(username=row[1])
                studentUser.set_password(password)
                studentUser.save()
                student_group = Group.objects.get(name='Students')
                studentUser.groups.add(student_group)
                studentUser.save()

                #getting data from db for foreign keys                
                parentData = ParentsInfo.objects.all()
                for tempparent in parentData:
                    if encryptionHelper.decrypt(tempparent.email) == row[7]:
                        parent = tempparent
                  
                school = School.objects.filter(name__icontains=row[6]).first()                
                teacher = TeacherInCharge.objects.filter(user = request.user).first()
                #creating student                 
                dob = datetime.strptime(row[5],'%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
                student = StudentsInfo(name=row[0],address=row[2],rollno=row[3],gender=row[4],dob=dob,school=school,first_password=password,teacher=teacher)
                student.parent = parent
                student.user = studentUser
                student.save()    
        
        messages.success(request, 'Registration Completed')
        return redirect('/bulkRegister')


@login_required(login_url='/login')
@user_passes_test(is_teacher,login_url='/forbidden')
def getTemplate(request):
    output = io.BytesIO()
    
    wb =  xlsxwriter.Workbook(output)
    ws = wb.add_worksheet("Parents Data")
    ws2 = wb.add_worksheet("Students Data")
    
    columns = ["Parent Email","Parent Name","Gender","Age","Address","Pincode","No of family members","Children Count","City","State","Education","Occupation","Religion","Type of family"]
    columns2 = ["Student Name","Address","Registration No","Gender","DOB","School","Parents email"]    

    sampleParentData =["john@gmail.com","John Doe","Male","29","Mumbai","400001","5","2","Mumbai","Maharashtra","BTech","Engineer","Hindu","Nuclear"]


    sampleStudentData =["Jane Doe","Mumbai","1234","Female",date.today().strftime("%d-%m-%Y"),"K.J Somaiya School","john@gmail.com"]
    row_num = 0    

    
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num]) # at 0 row 0 column 

    for col_num in range(len(columns2)):
        ws2.write(row_num, col_num, columns2[col_num]) # at 0 row 0 column 

    row_num+=1
    for col_num in range(len(sampleParentData)):
        ws.write(row_num, col_num, sampleParentData[col_num])

    for col_num in range(len(sampleStudentData)):
        ws2.write(row_num, col_num, sampleStudentData[col_num])
    wb.close()

    # construct response
    output.seek(0)
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename=template.xlsx"    
    return response  

@login_required(login_url='/login')
@user_passes_test(is_teacher,login_url='/forbidden')
def downloadData(request):

    output = io.BytesIO()
    wb =  xlsxwriter.Workbook(output)
    parentSheet = wb.add_worksheet("Parents Data")
    studentSheet = wb.add_worksheet("Students Data")
    encryptionHelper = EncryptionHelper()

    #student sheet
    row_num = 0
    studentColumns = ['Name','Roll No','DOB','Gender','Address','School','Parent\'s Email','Username', 'Password']
    for col_num in range(len(studentColumns)):
        studentSheet.write(row_num, col_num, studentColumns[col_num]) # at 0 row 0 column 

    teacher = TeacherInCharge.objects.filter(user = request.user).first()
    rows = StudentsInfo.objects.filter(teacher=teacher).values_list('name','rollno','dob','gender','address','school','parent','user','first_password','password_changed')    
    parentEmail = []    
    for row in rows:        
        row_num += 1
        for col_num in range(len(studentColumns)):            
            if(col_num==0):                                               
                studentSheet.write(row_num, col_num, encryptionHelper.decrypt(row[col_num]))
            
            elif(col_num==2):
                studentSheet.write(row_num, col_num, row[col_num].strftime('%d/%b/%Y'))

            elif(col_num==5):
                school = School.objects.get(pk=row[col_num])
                studentSheet.write(row_num, col_num, school.name)

            elif(col_num==6):
                parent = ParentsInfo.objects.get(pk=row[col_num])
                email = encryptionHelper.decrypt(parent.email)
                if email not in parentEmail:
                    parentEmail.append(email)
                studentSheet.write(row_num, col_num, email)

            elif(col_num==7):
                user = User.objects.get(pk=row[col_num])
                studentSheet.write(row_num, col_num, user.username)

            elif(col_num==8):
                msg = row[col_num]
                if row[col_num+1]:
                    msg = "Already Changed"
                studentSheet.write(row_num, col_num, msg)

            elif col_num==9:
                continue

            else:
                studentSheet.write(row_num, col_num, row[col_num])

    
    # Sheet header, first row
    row_num = 0
    
    #parent sheet
    parentColumns = ['Name','Email','Age','Gender','Address','City','State','Pincode','Education','Occupation','Religion','Family count','Children count','Family Type','Username','Password']
    for col_num in range(len(parentColumns)):
        parentSheet.write(row_num, col_num, parentColumns[col_num]) # at 0 row 0 column 

    rows = ParentsInfo.objects.all().values_list('name','email','age','gender','address','city','state','pincode','edu','occupation','religion','no_of_family_members','children_count','type_of_family','user','first_password','password_changed')    
    for row in rows:                
        if encryptionHelper.decrypt(row[1]) in parentEmail:
            row_num += 1
            for col_num in range(len(parentColumns)):   
                               
                if(col_num==0):              
                    parentSheet.write(row_num, col_num, encryptionHelper.decrypt(row[col_num]))
                
                elif(col_num==1):             
                    parentSheet.write(row_num, col_num, encryptionHelper.decrypt(row[col_num]))

                elif(col_num==5):
                    city = City.objects.get(pk=row[col_num])
                    parentSheet.write(row_num, col_num, city.city)

                elif(col_num==6):
                    state = State.objects.get(pk=row[col_num])
                    parentSheet.write(row_num, col_num, state.state)

                elif(col_num==8):
                    education = Education.objects.get(pk=row[col_num])
                    parentSheet.write(row_num, col_num, education.education)

                elif(col_num==9):
                    occupation = Occupation.objects.get(pk=row[col_num])
                    parentSheet.write(row_num, col_num, occupation.occupation)

                elif(col_num==10):
                    religion = ReligiousBelief.objects.get(pk=row[col_num])
                    parentSheet.write(row_num, col_num, religion.religion)

                elif(col_num==13):
                    familyType = FamilyType.objects.get(pk=row[col_num])
                    parentSheet.write(row_num, col_num, familyType.family)

                elif(col_num==14):
                    user = User.objects.get(pk=row[col_num])
                    parentSheet.write(row_num, col_num, user.username)

                elif(col_num==15):
                    msg = row[col_num]
                    if row[col_num+1]:
                        msg = "Already Changed"
                    parentSheet.write(row_num, col_num, msg)

                elif col_num==16:
                    continue

                else:
                    parentSheet.write(row_num, col_num, row[col_num])

    
    wb.close()
    # construct response
    output.seek(0)
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename=Parent and Student list.xlsx"
    return response


# def getFirstModule(request):
#     if(request.method == "GET"):
#         form = FirstModuleForm()
#         return render(request,'registration_form/first_module.html',{'form':form})
#     else:
#         request.session['data'] = request.POST
#         return redirect('/nutriPartTwo')


# def nutri(request):    
#     return redirect('nutriPartTwo')

# def nutriPartTwo(request):
#     if(request.method == "GET"):
#         form = FirstModuleForm()
#         return render(request,'registration_form/first_module_second.html',{'form':form})


@login_required(login_url='/login')
@user_passes_test(is_student,login_url='/forbidden')
def student_dashboard(request):
    return render(request,'registration_form/student_dashboard.html')


@login_required(login_url='/login')
@user_passes_test(is_teacher,login_url='/forbidden')
def teacher_dashboard(request):
    teacher = TeacherInCharge.objects.get(user=request.user)
    total_students = teacher.studentsinfo_set.all()
    results = []
    closed_sessions = FormDetails.objects.filter(teacher=teacher,open=False)

    for session in closed_sessions:
        temp_list = [session.form,session.start_timestamp,session.end_timestamp]
        if session.pre:
            temp_list.append("Pre Test")
        else:
            temp_list.append("Post Test")
        count = 0
        for student in total_students:
            if ModuleOne.objects.filter(student=student,submission_timestamp__gte=session.start_timestamp,submission_timestamp__lte=session.end_timestamp).exists():
                draftForm = ModuleOne.objects.filter(student=student,submission_timestamp__gte=session.start_timestamp,submission_timestamp__lte=session.end_timestamp).first()
                if not draftForm.draft:
                    count += 1
                    
        temp_list.append(count)
        temp_list.append(len(total_students))
        temp_list.append(session.id)
        results.append(temp_list)

    open_sessions = FormDetails.objects.filter(teacher=teacher,open=True)
    results2 = []    
    for session in open_sessions:
        temp_list = [session.form,session.start_timestamp]
        if session.pre:
            temp_list.append("Pre Test")
        else:
            temp_list.append("Post Test")
        count = 0
        for student in total_students:
            if ModuleOne.objects.filter(student=student,submission_timestamp__gte=session.start_timestamp).exists():
                draftForm = ModuleOne.objects.filter(student=student,submission_timestamp__gte=session.start_timestamp).first()
                if not draftForm.draft:
                    count += 1
                    
        temp_list.append(count)
        temp_list.append(len(total_students))
        temp_list.append(session.id)
        results2.append(temp_list)

    return render(request,'registration_form/teacher_dashboard.html',{'results':results,'results2':results2})


def createTempDict(postData):
    temp = {}       
    for key in postData:                            
            if key == 'source_fruits_vegetables' or key=='grow_own_food':
                temp[key] = postData.getlist(key)
            else:
                temp[key] = postData[key]
    del temp['csrfmiddlewaretoken'] 
    return temp


def creatingOrUpdatingDrafts(temp,user):
    student = StudentsInfo.objects.get(user=user)
    startdate = FormDetails.objects.get(form=Form.objects.get(name='moduleOne'),teacher=student.teacher,open=True).start_timestamp        
    if ModuleOne.objects.filter(student=student,submission_timestamp__gte=startdate).exists(): 
        draftForm = ModuleOne.objects.get(student=StudentsInfo.objects.get(user=user),submission_timestamp__gte=startdate)  
        if draftForm.draft:
        # updating drafts
            for name in ModuleOne._meta.get_fields():
                name = name.column
                if name == 'id' or name == 'student_id' or name == 'draft':
                        continue
                if name in temp:
                    setattr(draftForm, name, temp[name])      
                else:
                    setattr(draftForm, name, getattr(draftForm, name) or None) 

            draftForm.submission_timestamp = datetime.datetime.now()
            draftForm.save()
            return True
    else:
        return False

@login_required(login_url='/login')
@user_passes_test(is_student,login_url='/forbidden')
def draft(request):
    if 'parent_dashboard' in request.META.get('HTTP_REFERER').split('/'):
        module = request.META.get('HTTP_REFERER').split('/')[-1]
        id =  request.META.get('HTTP_REFERER').split('/')[-2]    
        user = StudentsInfo.objects.get(pk=id).user   
    else:
        module = request.META.get('HTTP_REFERER').split('/')[-2]  
        user = request.user  

    #1st Page
    if module=="moduleOne":         
        #for removing csrf field 
        temp = createTempDict(request.POST)  
        #checking if draft exists    
        if not creatingOrUpdatingDrafts(temp,user):
            #creating new record
            form = ModuleOne(**temp)
            form.student = StudentsInfo.objects.get(user=user)
            form.draft = True
            formType = getFormType('moduleOne')
            form.pre = 1 if formType=='PreTest' else 0
            form.submission_timestamp = datetime.datetime.now()
            form.save()
    #2nd Page
    elif module=="moduleOne-2":
        temp = createTempDict(request.POST)  
        creatingOrUpdatingDrafts(temp,user)
    #3rd Page
    elif module=="moduleOne-3":
        temp = createTempDict(request.POST)             
        creatingOrUpdatingDrafts(temp,user)
    return redirect(request.META.get('HTTP_REFERER'))
 

def getFormType(moduleType):
    module = Form.objects.get(name=moduleType)
    formType = FormDetails.objects.filter(form=module,open=True).first()
    return 'PreTest' if formType.pre else 'PostTest'

@login_required(login_url='/login')
@user_passes_test(is_student,login_url='/forbidden')
@isActive('moduleOne','student')
def moduleOne(request,user=None):
    if(request.method=="GET"):                      
        if user==None:
            user = request.user

        student = StudentsInfo.objects.get(user=user)
        startdate = FormDetails.objects.get(form=Form.objects.get(name='moduleOne'),teacher=student.teacher,open=True).start_timestamp        
        if ModuleOne.objects.filter(student=student,submission_timestamp__gte=startdate).exists(): 
            draftForm = ModuleOne.objects.get(student=StudentsInfo.objects.get(user=user),submission_timestamp__gte=startdate)            
            if draftForm.draft:
                mod = ModuleOneForm()
                temp = {}
                for name in ModuleOne._meta.get_fields():
                    name = name.column
                    if name in mod.fields:                   
                        if name == 'source_fruits_vegetables' or name=='grow_own_food':
                            temp[name] = ast.literal_eval(getattr(draftForm, name) or '[]')
                        else:
                            temp[name] = getattr(draftForm, name)
                                    
                form = ModuleOneForm(temp)   
                formPre = getFormType('moduleOne')        
                return render(request,'registration_form/module_one.html',{'form':form,'formPre':formPre})
            else:
                return redirect('/already_filled')
                        
        #new form
        else:            
            form = ModuleOneForm()
            formPre = getFormType('moduleOne')
            return render(request,'registration_form/module_one.html',{'form':form,'formPre':formPre})

    #POST            
    else:  
                
        flag = False
        if user == None:
            flag = True   
            user = request.user  
        form = ModuleOneForm(request.POST)             
        
        if form.is_valid():
            temp = createTempDict(request.POST) 

            if not creatingOrUpdatingDrafts(temp,user):
                #creating new record        
                form = ModuleOne(**temp)
                form.student = StudentsInfo.objects.get(user=user)
                form.draft = True
                formType = getFormType('moduleOne')
                form.pre = 1 if formType=='PreTest' else 0
                form.submission_timestamp = datetime.datetime.now()
                form.save()

            if flag:
                return redirect('/moduleOne-2')
            else:                
                return redirect('parentsModuleOne2',id=StudentsInfo.objects.get(user=user).id)
        
        else:            
            formPre = getFormType('moduleOne')
            return render(request,'registration_form/module_one.html',{'form':form,'formPre':formPre})    

@login_required(login_url='/login')     
@user_passes_test(is_student,login_url='/forbidden')
@isActive('moduleOne','student')
def moduleOne2(request,user=None):
    if(request.method=="GET"):
        if user==None:
            user = request.user
            
        student = StudentsInfo.objects.get(user=user)
        startdate = FormDetails.objects.get(form=Form.objects.get(name='moduleOne'),teacher=student.teacher,open=True).start_timestamp        
        if ModuleOne.objects.filter(student=student,submission_timestamp__gte=startdate).exists(): 
            draftForm = ModuleOne.objects.get(student=StudentsInfo.objects.get(user=user),submission_timestamp__gte=startdate)            
            if draftForm.draft:
                mod = ModuleOneForm2()
                temp = {}
                for name in ModuleOne._meta.get_fields():
                    name = name.column
                    if name in mod.fields:  
                        temp[name] = getattr(draftForm, name) or None
        
                form = ModuleOneForm2(temp)                           
                formPre = getFormType('moduleOne')                
                return render(request,'registration_form/module_one2.html',{'form':form,'formPre':formPre})
            else:
                return redirect('/already_filled')
          
        #new form
        else:            
            form = ModuleOneForm2()
            formPre = getFormType('moduleOne')            
            return render(request,'registration_form/module_one2.html',{'form':form,'formPre':formPre})
    #POST
    else:
        flag = False
        if user == None:
            flag = True   
            user = request.user 
        form = ModuleOneForm2(request.POST)                

        if form.is_valid():
            temp = createTempDict(request.POST)             
            creatingOrUpdatingDrafts(temp,user)

            if flag:
                return redirect('/moduleOne-3')
            else:                
                return redirect('parentsModuleOne3',id=StudentsInfo.objects.get(user=user).id)                    
        else:
            formPre = getFormType('moduleOne')
            return render(request,'registration_form/module_one2.html',{'form':form,'formPre':'formPre'})


@login_required(login_url='/login')            
@user_passes_test(is_student,login_url='/forbidden')
@isActive('moduleOne','student')
def moduleOne3(request,user=None):
    if(request.method=="GET"):
        if user==None:
            user = request.user

        student = StudentsInfo.objects.get(user=user)
        startdate = FormDetails.objects.get(form=Form.objects.get(name='moduleOne'),teacher=student.teacher,open=True).start_timestamp        
        if ModuleOne.objects.filter(student=student,submission_timestamp__gte=startdate).exists(): 
            draftForm = ModuleOne.objects.get(student=StudentsInfo.objects.get(user=user),submission_timestamp__gte=startdate)            
            if draftForm.draft:
                mod = ModuleOneForm3()
                temp = {}
                for name in ModuleOne._meta.get_fields():
                    name = name.column
                    if name in mod.fields:                  
                        temp[name] = getattr(draftForm, name) or None
                form = ModuleOneForm3(temp)                           
                formPre = getFormType('moduleOne')
                return render(request,'registration_form/module_one3.html',{'form':form,'formPre':formPre})
            else:
                return redirect('/already_filled')
        #new form
        else:            
            form = ModuleOneForm3()
            formPre = getFormType('moduleOne')
            return render(request,'registration_form/module_one3.html',{'form':form,'formPre':formPre})
    #POST
    else:
        flag = False
        if user == None:
            flag = True   
            user = request.user
        form = ModuleOneForm3(request.POST)                
            
        #valid form
        if form.is_valid():                    
            temp = createTempDict(request.POST)
            student = StudentsInfo.objects.get(user=user)
            startdate = FormDetails.objects.get(form=Form.objects.get(name='moduleOne'),teacher=student.teacher,open=True).start_timestamp        
            draftForm = ModuleOne.objects.get(student=StudentsInfo.objects.get(user=user),submission_timestamp__gte=startdate)
            if draftForm.draft:
                for name in ModuleOne._meta.get_fields():
                    name = name.column
                    if name == 'id' or name == 'student_id' or name == 'draft':
                            continue                    
                    elif name == 'source_fruits_vegetables' or name == 'grow_own_food':
                        list = '; '.join(ast.literal_eval(getattr(draftForm, name)))
                        setattr(draftForm, name, list)      
                    elif name in temp:
                        setattr(draftForm, name, temp[name])      

                draftForm.draft = False
                draftForm.submission_timestamp = datetime.datetime.now()
                draftForm.save()
                if flag:
                    return redirect('/student_dashboard')
                else:
                    return redirect('/parent_dashboard')            
        #invalid form
        else:                               
            formPre = getFormType('moduleOne')
            return render(request,'registration_form/module_one3.html',{'form':form,'formPre':formPre})
            

def forbidden(request):
    raise PermissionDenied

@login_required(login_url='/login')
@user_passes_test(is_parent,login_url='/forbidden')
def showStudent(request,id):
    student = StudentsInfo.objects.get(pk=id)
    encryptionHelper = EncryptionHelper()
    return render(request,'registration_form/student_modules.html')

@login_required(login_url='/login')
@user_passes_test(is_parent,login_url='/forbidden')
@isActive('moduleOne','parent')
def parentModuleOne(request,id):
    user = StudentsInfo.objects.get(pk=id).user
    return moduleOne(request,user)

@login_required(login_url='/login')    
@user_passes_test(is_parent,login_url='/forbidden')
@isActive('moduleOne','parent')
def parentModuleOne2(request,id):
    user = StudentsInfo.objects.get(pk=id).user
    return moduleOne2(request,user)

@login_required(login_url='/login')  
@user_passes_test(is_parent,login_url='/forbidden')
@isActive('moduleOne','parent')
def parentModuleOne3(request,id):
    user = StudentsInfo.objects.get(pk=id).user
    return moduleOne3(request,user)

@login_required(login_url='/login')
@user_passes_test(is_student,login_url='/forbidden')
def previous(request):
    link = request.META.get('HTTP_REFERER').split('/')    
    if 'parent_dashboard' in link:
        if link[-1] == 'moduleOne-2':
            link[-1] = 'moduleOne'
        elif link[-1] == 'moduleOne-3':
            link[-1] = 'moduleOne-2' 
    else:
        if link[-2] == 'moduleOne-2':
            link[-2] = 'moduleOne'
        elif link[-2] == 'moduleOne-3':
            link[-2] = 'moduleOne-2' 
    newLink = '/'.join(link)
    return redirect(newLink)

@login_required(login_url='/login')
@user_passes_test(is_teacher,login_url='/forbidden')
def manageForms(request):
    if request.method == "GET":
        moduleOne={}
        moduleTwo={}
        moduleThree={}
        form = Form.objects.get(name='moduleOne')
        teacher = TeacherInCharge.objects.get(user=request.user)
        if FormDetails.objects.filter(form=form,teacher=teacher).exists():

            form = FormDetails.objects.filter(form=form,teacher=teacher,open=True).order_by('-start_timestamp').first()            
            if form:
                if form.pre:
                    moduleOne['pre'] = True
                    moduleOne['post'] = False

                else:
                    moduleOne['post'] = True
                    moduleOne['pre'] = False

        return render(request,'registration_form/manage_forms_teacher.html',{'moduleOne': moduleOne, 'moduleTwo':moduleTwo, 'moduleThree':moduleThree})
    else:
        
        if 'moduleOne' in request.POST:
            module_one_pre = request.POST.get('module_one_pre', False)
            module_one_post = request.POST.get('module_one_post', False)            
            if module_one_pre == 'on' and module_one_post == 'on':                
                messages.error(request, 'Cannot select both PreTest and PostTest')
                return redirect('/manage-forms')


            form = Form.objects.get(name='moduleOne')
            teacher = TeacherInCharge.objects.get(user=request.user)

            if module_one_pre=='on':
                if not FormDetails.objects.filter(form=form,teacher=teacher,pre=True,open=True).exists():
                    update = FormDetails(form=form,teacher=teacher,pre=True,open=True,start_timestamp=datetime.datetime.now())
                    update.save()
            else:                
                if FormDetails.objects.filter(form=form,teacher=teacher,pre=True,open=True).exists():                    
                    update = FormDetails.objects.filter(form=form,teacher=teacher,pre=True,open=True).first()                    
                    update.open = False
                    update.end_timestamp = datetime.datetime.now()
                    teacher = TeacherInCharge.objects.get(user=request.user)
                    total_students = teacher.studentsinfo_set.all()
                    for student in total_students:
                        if ModuleOne.objects.filter(student=student,submission_timestamp__gte=update.start_timestamp,submission_timestamp__lte=update.end_timestamp,draft=True,pre=True).exists():
                            draftForm = ModuleOne.objects.filter(student=student,submission_timestamp__gte=update.start_timestamp,submission_timestamp__lte=update.end_timestamp,draft=True,pre=True).first()
                            draftForm.delete()
                    update.save()

            if module_one_post=='on':
                if not FormDetails.objects.filter(form=form,teacher=teacher,pre=False,open=True).exists():
                    update = FormDetails(form=form,teacher=teacher,pre=False,open=True,start_timestamp=datetime.datetime.now())
                    update.save()
            else:                            
                if FormDetails.objects.filter(form=form,teacher=teacher,pre=False,open=True).exists():                    
                    update = FormDetails.objects.filter(form=form,teacher=teacher,pre=False,open=True).first()                    
                    update.open = False
                    update.end_timestamp = datetime.datetime.now()
                    teacher = TeacherInCharge.objects.get(user=request.user)
                    total_students = teacher.studentsinfo_set.all()
                    for student in total_students:
                        if ModuleOne.objects.filter(student=student,submission_timestamp__gte=update.start_timestamp,submission_timestamp__lte=update.end_timestamp,draft=True,pre=False).exists():
                            draftForm = ModuleOne.objects.filter(student=student,submission_timestamp__gte=update.start_timestamp,submission_timestamp__lte=update.end_timestamp,draft=True,pre=False).first()
                            draftForm.delete()
                    update.save()

        
        module_two_pre = request.POST.get('module_two_pre', False)
        module_two_post = request.POST.get('module_two_post', False)
        module_three_pre = request.POST.get('module_three_pre', False)
        module_three_post = request.POST.get('module_three_post', False)
            
        # if module_one_pre == module_one_post or module_two_pre == module_two_post or module_three_pre == module_three_post:
        # FormDetails.objects.filter(form= )
        return redirect('/manage-forms')

@login_required(login_url='/login')
@user_passes_test(is_teacher,login_url='/forbidden')
def getFormDetails(request,id):
    session = FormDetails.objects.get(pk=id)
    print(session)
    teacher = session.teacher
    total_students = teacher.studentsinfo_set.all()
    
    filled_students = []
    not_filled_students = []
    encryptionHelper = EncryptionHelper()
    open = True
    if not session.open:
        open = False
        temp_list = [session.form,session.start_timestamp,session.end_timestamp]
    else:
        temp_list = [session.form,session.start_timestamp]

    if session.pre:
        temp_list.append("Pre Test")
    else:
        temp_list.append("Post Test")
    count = 0
    for student in total_students:
        temp = []
        if not session.open:
            if ModuleOne.objects.filter(student=student,submission_timestamp__gte=session.start_timestamp,submission_timestamp__lte=session.end_timestamp).exists():
                draftForm = ModuleOne.objects.filter(student=student,submission_timestamp__gte=session.start_timestamp,submission_timestamp__lte=session.end_timestamp).first()            
                
                if not draftForm.draft:
                    count += 1
                    temp.append(encryptionHelper.decrypt(student.name))
                    temp.append(draftForm.submission_timestamp)
                    filled_students.append(temp)

                else:
                    temp.append(encryptionHelper.decrypt(student.name))
                    temp.append('-')
                    not_filled_students.append(temp)
            else:
                temp.append(encryptionHelper.decrypt(student.name))
                temp.append('-')
                not_filled_students.append(temp)

        else:
            if ModuleOne.objects.filter(student=student,submission_timestamp__gte=session.start_timestamp).exists():
                draftForm = ModuleOne.objects.filter(student=student,submission_timestamp__gte=session.start_timestamp).first()            
            
                if not draftForm.draft:
                    count += 1
                    temp.append(encryptionHelper.decrypt(student.name))
                    temp.append(draftForm.submission_timestamp)
                    filled_students.append(temp)

                else:
                    temp.append(encryptionHelper.decrypt(student.name))
                    temp.append('-')
                    not_filled_students.append(temp)
            else:
                temp.append(encryptionHelper.decrypt(student.name))
                temp.append('-')
                not_filled_students.append(temp)

                    
    temp_list.append(count)
    temp_list.append(len(total_students))    
    
    return render(request,'registration_form/teacher_dashboard_getDetails.html',{'result':temp_list,'filled_students':filled_students,'not_filled_students':not_filled_students,'open':open})





def uploadData(request):

    state_arr  = ["Andaman & Nicobar","Andhra Pradesh","Arunachal Pradesh","Assam","Bihar","Chandigarh","Chhattisgarh","Dadra & Nagar Haveli","Daman & Diu","Delhi","Goa","Gujarat","Haryana","Himachal Pradesh","Jammu & Kashmir","Jharkhand","Karnataka","Kerala","Lakshadweep","Madhya Pradesh","Maharashtra","Manipur","Meghalaya","Mizoram","Nagaland","Orissa","Pondicherry","Punjab","Rajasthan","Sikkim","Tamil Nadu","Tripura","Uttar Pradesh","Uttaranchal","West Bengal"]

    s_a = []
    s_a.append("Alipur | Andaman Island | Anderson Island | Arainj-Laka-Punga | Austinabad | Bamboo Flat | Barren Island | Beadonabad | Betapur | Bindraban | Bonington | Brookesabad | Cadell Point | Calicut | Chetamale | Cinque Islands | Defence Island | Digilpur | Dolyganj | Flat Island | Geinyale | Great Coco Island | Haddo | Havelock Island | Henry Lawrence Island | Herbertabad | Hobdaypur | Ilichar | Ingoie | Inteview Island | Jangli Ghat | Jhon Lawrence Island | Karen | Kartara | KYD Islannd | Landfall Island | Little Andmand | Little Coco Island | Long Island | Maimyo | Malappuram | Manglutan | Manpur | Mitha Khari | Neill Island | Nicobar Island | North Brother Island | North Passage Island | North Sentinel Island | Nothen Reef Island | Outram Island | Pahlagaon | Palalankwe | Passage Island | Phaiapong | Phoenix Island | Port Blair | Preparis Island | Protheroepur | Rangachang | Rongat | Rutland Island | Sabari | Saddle Peak | Shadipur | Smith Island | Sound Island | South Sentinel Island | Spike Island | Tarmugli Island | Taylerabad | Titaije | Toibalawe | Tusonabad | West Island | Wimberleyganj | Yadita")
    s_a.append("Achampet | Adilabad | Adoni | Alampur | Allagadda | Alur | Amalapuram | Amangallu | Anakapalle | Anantapur | Andole | Araku | Armoor | Asifabad | Aswaraopet | Atmakur | B. Kothakota | Badvel | Banaganapalle | Bandar | Bangarupalem | Banswada | Bapatla | Bellampalli | Bhadrachalam | Bhainsa | Bheemunipatnam | Bhimadole | Bhimavaram | Bhongir | Bhooragamphad | Boath | Bobbili | Bodhan | Chandoor | Chavitidibbalu | Chejerla | Chepurupalli | Cherial | Chevella | Chinnor | Chintalapudi | Chintapalle | Chirala | Chittoor | Chodavaram | Cuddapah | Cumbum | Darsi | Devarakonda | Dharmavaram | Dichpalli | Divi | Donakonda | Dronachalam | East Godavari | Eluru | Eturnagaram | Gadwal | Gajapathinagaram | Gajwel | Garladinne | Giddalur | Godavari | Gooty | Gudivada | Gudur | Guntur | Hindupur | Hunsabad | Huzurabad | Huzurnagar | Hyderabad | Ibrahimpatnam | Jaggayyapet | Jagtial | Jammalamadugu | Jangaon | Jangareddygudem | Jannaram | Kadiri | Kaikaluru | Kakinada | Kalwakurthy | Kalyandurg | Kamalapuram | Kamareddy | Kambadur | Kanaganapalle | Kandukuru | Kanigiri | Karimnagar | Kavali | Khammam | Khanapur (AP) | Kodangal | Koduru | Koilkuntla | Kollapur | Kothagudem | Kovvur | Krishna | Krosuru | Kuppam | Kurnool | Lakkireddipalli | Madakasira | Madanapalli | Madhira | Madnur | Mahabubabad | Mahabubnagar | Mahadevapur | Makthal | Mancherial | Mandapeta | Mangalagiri | Manthani | Markapur | Marturu | Medachal | Medak | Medarmetla | Metpalli | Mriyalguda | Mulug | Mylavaram | Nagarkurnool | Nalgonda | Nallacheruvu | Nampalle | Nandigama | Nandikotkur | Nandyal | Narasampet | Narasaraopet | Narayanakhed | Narayanpet | Narsapur | Narsipatnam | Nazvidu | Nelloe | Nellore | Nidamanur | Nirmal | Nizamabad | Nuguru | Ongole | Outsarangapalle | Paderu | Pakala | Palakonda | Paland | Palmaneru | Pamuru | Pargi | Parkal | Parvathipuram | Pathapatnam | Pattikonda | Peapalle | Peddapalli | Peddapuram | Penukonda | Piduguralla | Piler | Pithapuram | Podili | Polavaram | Prakasam | Proddatur | Pulivendla | Punganur | Putturu | Rajahmundri | Rajampeta | Ramachandrapuram | Ramannapet | Rampachodavaram | Rangareddy | Rapur | Rayachoti | Rayadurg | Razole | Repalle | Saluru | Sangareddy | Sathupalli | Sattenapalle | Satyavedu | Shadnagar | Siddavattam | Siddipet | Sileru | Sircilla | Sirpur Kagaznagar | Sodam | Sompeta | Srikakulam | Srikalahasthi | Srisailam | Srungavarapukota | Sudhimalla | Sullarpet | Tadepalligudem | Tadipatri | Tanduru | Tanuku | Tekkali | Tenali | Thungaturthy | Tirivuru | Tirupathi | Tuni | Udaygiri | Ulvapadu | Uravakonda | Utnor | V.R. Puram | Vaimpalli | Vayalpad | Venkatgiri | Venkatgirikota | Vijayawada | Vikrabad | Vinjamuru | Vinukonda | Visakhapatnam | Vizayanagaram | Vizianagaram | Vuyyuru | Wanaparthy | Warangal | Wardhannapet | Yelamanchili | Yelavaram | Yeleswaram | Yellandu | Yellanuru | Yellareddy | Yerragondapalem | Zahirabad")
    s_a.append("Along | Anini | Anjaw | Bameng | Basar | Changlang | Chowkhem | Daporizo | Dibang Valley | Dirang | Hayuliang | Huri | Itanagar | Jairampur | Kalaktung | Kameng | Khonsa | Kolaring | Kurung Kumey | Lohit | Lower Dibang Valley | Lower Subansiri | Mariyang | Mechuka | Miao | Nefra | Pakkekesang | Pangin | Papum Pare | Passighat | Roing | Sagalee | Seppa | Siang | Tali | Taliha | Tawang | Tezu | Tirap | Tuting | Upper Siang | Upper Subansiri | Yiang Kiag")
    s_a.append("Abhayapuri | Baithalangshu | Barama | Barpeta Road | Bihupuria | Bijni | Bilasipara | Bokajan | Bokakhat | Boko | Bongaigaon | Cachar | Cachar Hills | Darrang | Dhakuakhana | Dhemaji | Dhubri | Dibrugarh | Digboi | Diphu | Goalpara | Gohpur | Golaghat | Guwahati | Hailakandi | Hajo | Halflong | Hojai | Howraghat | Jorhat | Kamrup | Karbi Anglong | Karimganj | Kokarajhar | Kokrajhar | Lakhimpur | Maibong | Majuli | Mangaldoi | Mariani | Marigaon | Moranhat | Morigaon | Nagaon | Nalbari | Rangapara | Sadiya | Sibsagar | Silchar | Sivasagar | Sonitpur | Tarabarihat | Tezpur | Tinsukia | Udalgiri | Udalguri | UdarbondhBarpeta")
    s_a.append("Adhaura | Amarpur | Araria | Areraj | Arrah | Arwal | Aurangabad | Bagaha | Banka | Banmankhi | Barachakia | Barauni | Barh | Barosi | Begusarai | Benipatti | Benipur | Bettiah | Bhabhua | Bhagalpur | Bhojpur | Bidupur | Biharsharif | Bikram | Bikramganj | Birpur | Buxar | Chakai | Champaran | Chapara | Dalsinghsarai | Danapur | Darbhanga | Daudnagar | Dhaka | Dhamdaha | Dumraon | Ekma | Forbesganj | Gaya | Gogri | Gopalganj | H.Kharagpur | Hajipur | Hathua | Hilsa | Imamganj | Jahanabad | Jainagar | Jamshedpur | Jamui | Jehanabad | Jhajha | Jhanjharpur | Kahalgaon | Kaimur (Bhabua) | Katihar | Katoria | Khagaria | Kishanganj | Korha | Lakhisarai | Madhepura | Madhubani | Maharajganj | Mahua | Mairwa | Mallehpur | Masrakh | Mohania | Monghyr | Motihari | Motipur | Munger | Muzaffarpur | Nabinagar | Nalanda | Narkatiaganj | Naugachia | Nawada | Pakribarwan | Pakridayal | Patna | Phulparas | Piro | Pupri | Purena | Purnia | Rafiganj | Rajauli | Ramnagar | Raniganj | Raxaul | Rohtas | Rosera | S.Bakhtiarpur | Saharsa | Samastipur | Saran | Sasaram | Seikhpura | Sheikhpura | Sheohar | Sherghati | Sidhawalia | Singhwara | Sitamarhi | Siwan | Sonepur | Supaul | Thakurganj | Triveniganj | Udakishanganj | Vaishali | Wazirganj")
    s_a.append("Chandigarh | Mani Marja")
    s_a.append("Ambikapur | Antagarh | Arang | Bacheli | Bagbahera | Bagicha | Baikunthpur | Balod | Balodabazar | Balrampur | Barpalli | Basana | Bastanar | Bastar | Bderajpur | Bemetara | Berla | Bhairongarh | Bhanupratappur | Bharathpur | Bhatapara | Bhilai | Bhilaigarh | Bhopalpatnam | Bijapur | Bilaspur | Bodla | Bokaband | Chandipara | Chhinagarh | Chhuriakala | Chingmut | Chuikhadan | Dabhara | Dallirajhara | Dantewada | Deobhog | Dhamda | Dhamtari | Dharamjaigarh | Dongargarh | Durg | Durgakondal | Fingeshwar | Gariaband | Garpa | Gharghoda | Gogunda | Ilamidi | Jagdalpur | Janjgir | Janjgir-Champa | Jarwa | Jashpur | Jashpurnagar | Kabirdham-Kawardha | Kanker | Kasdol | Kathdol | Kathghora | Kawardha | Keskal | Khairgarh | Kondagaon | Konta | Korba | Korea | Kota | Koyelibeda | Kuakunda | Kunkuri | Kurud | Lohadigundah | Lormi | Luckwada | Mahasamund | Makodi | Manendragarh | Manpur | Marwahi | Mohla | Mungeli | Nagri | Narainpur | Narayanpur | Neora | Netanar | Odgi | Padamkot | Pakhanjur | Pali | Pandaria | Pandishankar | Parasgaon | Pasan | Patan | Pathalgaon | Pendra | Pratappur | Premnagar | Raigarh | Raipur | Rajnandgaon | Rajpur | Ramchandrapur | Saraipali | Saranggarh | Sarona | Semaria | Shakti | Sitapur | Sukma | Surajpur | Surguja | Tapkara | Toynar | Udaipur | Uproda | Wadrainagar")
    s_a.append("Amal | Amli | Bedpa | Chikhli | Dadra & Nagar Haveli | Dahikhed | Dolara | Galonda | Kanadi | Karchond | Khadoli | Kharadpada | Kherabari | Kherdi | Kothar | Luari | Mashat | Rakholi | Rudana | Saili | Sili | Silvassa | Sindavni | Udva | Umbarkoi | Vansda | Vasona | Velugam")
    s_a.append("Brancavare | Dagasi | Daman | Diu | Magarvara | Nagwa | Pariali | Passo Covo")
    s_a.append("Central Delhi | East Delhi | New Delhi | North Delhi | North East Delhi | North West Delhi | South Delhi | South West Delhi | West Delhi")
    s_a.append("Canacona | Candolim | Chinchinim | Cortalim | Goa | Jua | Madgaon | Mahem | Mapuca | Marmagao | Panji | Ponda | Sanvordem | Terekhol")
    s_a.append("Ahmedabad | Ahwa | Amod | Amreli | Anand | Anjar | Ankaleshwar | Babra | Balasinor | Banaskantha | Bansada | Bardoli | Bareja | Baroda | Barwala | Bayad | Bhachav | Bhanvad | Bharuch | Bhavnagar | Bhiloda | Bhuj | Billimora | Borsad | Botad | Chanasma | Chhota Udaipur | Chotila | Dabhoi | Dahod | Damnagar | Dang | Danta | Dasada | Dediapada | Deesa | Dehgam | Deodar | Devgadhbaria | Dhandhuka | Dhanera | Dharampur | Dhari | Dholka | Dhoraji | Dhrangadhra | Dhrol | Dwarka | Fortsongadh | Gadhada | Gandhi Nagar | Gariadhar | Godhra | Gogodar | Gondal | Halol | Halvad | Harij | Himatnagar | Idar | Jambusar | Jamjodhpur | Jamkalyanpur | Jamnagar | Jasdan | Jetpur | Jhagadia | Jhalod | Jodia | Junagadh | Junagarh | Kalawad | Kalol | Kapad Wanj | Keshod | Khambat | Khambhalia | Khavda | Kheda | Khedbrahma | Kheralu | Kodinar | Kotdasanghani | Kunkawav | Kutch | Kutchmandvi | Kutiyana | Lakhpat | Lakhtar | Lalpur | Limbdi | Limkheda | Lunavada | M.M.Mangrol | Mahuva | Malia-Hatina | Maliya | Malpur | Manavadar | Mandvi | Mangrol | Mehmedabad | Mehsana | Miyagam | Modasa | Morvi | Muli | Mundra | Nadiad | Nakhatrana | Nalia | Narmada | Naswadi | Navasari | Nizar | Okha | Paddhari | Padra | Palanpur | Palitana | Panchmahals | Patan | Pavijetpur | Porbandar | Prantij | Radhanpur | Rahpar | Rajaula | Rajkot | Rajpipla | Ranavav | Sabarkantha | Sanand | Sankheda | Santalpur | Santrampur | Savarkundla | Savli | Sayan | Sayla | Shehra | Sidhpur | Sihor | Sojitra | Sumrasar | Surat | Surendranagar | Talaja | Thara | Tharad | Thasra | Una-Diu | Upleta | Vadgam | Vadodara | Valia | Vallabhipur | Valod | Valsad | Vanthali | Vapi | Vav | Veraval | Vijapur | Viramgam | Visavadar | Visnagar | Vyara | Waghodia | Wankaner")
    s_a.append("Adampur Mandi | Ambala | Assandh | Bahadurgarh | Barara | Barwala | Bawal | Bawanikhera | Bhiwani | Charkhidadri | Cheeka | Chhachrauli | Dabwali | Ellenabad | Faridabad | Fatehabad | Ferojpur Jhirka | Gharaunda | Gohana | Gurgaon | Hansi | Hisar | Jagadhari | Jatusana | Jhajjar | Jind | Julana | Kaithal | Kalanaur | Kalanwali | Kalka | Karnal | Kosli | Kurukshetra | Loharu | Mahendragarh | Meham | Mewat | Mohindergarh | Naraingarh | Narnaul | Narwana | Nilokheri | Nuh | Palwal | Panchkula | Panipat | Pehowa | Ratia | Rewari | Rohtak | Safidon | Sirsa | Siwani | Sonipat | Tohana | Tohsam | Yamunanagar")
    s_a.append("Amb | Arki | Banjar | Bharmour | Bilaspur | Chamba | Churah | Dalhousie | Dehra Gopipur | Hamirpur | Jogindernagar | Kalpa | Kangra | Kinnaur | Kullu | Lahaul | Mandi | Nahan | Nalagarh | Nirmand | Nurpur | Palampur | Pangi | Paonta | Pooh | Rajgarh | Rampur Bushahar | Rohru | Shimla | Sirmaur | Solan | Spiti | Sundernagar | Theog | Udaipur | Una")
    s_a.append("Akhnoor | Anantnag | Badgam | Bandipur | Baramulla | Basholi | Bedarwah | Budgam | Doda | Gulmarg | Jammu | Kalakot | Kargil | Karnah | Kathua | Kishtwar | Kulgam | Kupwara | Leh | Mahore | Nagrota | Nobra | Nowshera | Nyoma | Padam | Pahalgam | Patnitop | Poonch | Pulwama | Rajouri | Ramban | Ramnagar | Reasi | Samba | Srinagar | Udhampur | Vaishno Devi")
    s_a.append("Bagodar | Baharagora | Balumath | Barhi | Barkagaon | Barwadih | Basia | Bermo | Bhandaria | Bhawanathpur | Bishrampur | Bokaro | Bolwa | Bundu | Chaibasa | Chainpur | Chakardharpur | Chandil | Chatra | Chavparan | Daltonganj | Deoghar | Dhanbad | Dumka | Dumri | Garhwa | Garu | Ghaghra | Ghatsila | Giridih | Godda | Gomia | Govindpur | Gumla | Hazaribagh | Hunterganj | Ichak | Itki | Jagarnathpur | Jamshedpur | Jamtara | Japla | Jharmundi | Jhinkpani | Jhumaritalaiya | Kathikund | Kharsawa | Khunti | Koderma | Kolebira | Latehar | Lohardaga | Madhupur | Mahagama | Maheshpur Raj | Mandar | Mandu | Manoharpur | Muri | Nagarutatri | Nala | Noamundi | Pakur | Palamu | Palkot | Patan | Rajdhanwar | Rajmahal | Ramgarh | Ranchi | Sahibganj | Saraikela | Simaria | Simdega | Singhbhum | Tisri | Torpa")
    s_a.append("Afzalpur | Ainapur | Aland | Alur | Anekal | Ankola | Arsikere | Athani | Aurad | Bableshwar | Badami | Bagalkot | Bagepalli | Bailhongal | Bangalore | Bangalore Rural | Bangarpet | Bantwal | Basavakalyan | Basavanabagewadi | Basavapatna | Belgaum | Bellary | Belthangady | Belur | Bhadravati | Bhalki | Bhatkal | Bidar | Bijapur | Biligi | Chadchan | Challakere | Chamrajnagar | Channagiri | Channapatna | Channarayapatna | Chickmagalur | Chikballapur | Chikkaballapur | Chikkanayakanahalli | Chikkodi | Chikmagalur | Chincholi | Chintamani | Chitradurga | Chittapur | Cowdahalli | Davanagere | Deodurga | Devangere | Devarahippargi | Dharwad | Doddaballapur | Gadag | Gangavathi | Gokak | Gowribdanpur | Gubbi | Gulbarga | Gundlupet | H.B.Halli | H.D. Kote | Haliyal | Hampi | Hangal | Harapanahalli | Hassan | Haveri | Hebri | Hirekerur | Hiriyur | Holalkere | Holenarsipur | Honnali | Honnavar | Hosadurga | Hosakote | Hosanagara | Hospet | Hubli | Hukkeri | Humnabad | Hungund | Hunsagi | Hunsur | Huvinahadagali | Indi | Jagalur | Jamkhandi | Jewargi | Joida | K.R. Nagar | Kadur | Kalghatagi | Kamalapur | Kanakapura | Kannada | Kargal | Karkala | Karwar | Khanapur | Kodagu | Kolar | Kollegal | Koppa | Koppal | Koratageri | Krishnarajapet | Kudligi | Kumta | Kundapur | Kundgol | Kunigal | Kurugodu | Kustagi | Lingsugur | Madikeri | Madugiri | Malavalli | Malur | Mandya | Mangalore | Manipal | Manvi | Mashal | Molkalmuru | Mudalgi | Muddebihal | Mudhol | Mudigere | Mulbagal | Mundagod | Mundargi | Murugod | Mysore | Nagamangala | Nanjangud | Nargund | Narsimrajapur | Navalgund | Nelamangala | Nimburga | Pandavapura | Pavagada | Puttur | Raibag | Raichur | Ramdurg | Ranebennur | Ron | Sagar | Sakleshpur | Salkani | Sandur | Saundatti | Savanur | Sedam | Shahapur | Shankarnarayana | Shikaripura | Shimoga | Shirahatti | Shorapur | Siddapur | Sidlaghatta | Sindagi | Sindhanur | Sira | Sirsi | Siruguppa | Somwarpet | Sorab | Sringeri | Sriniwaspur | Srirangapatna | Sullia | T. Narsipur | Tallak | Tarikere | Telgi | Thirthahalli | Tiptur | Tumkur | Turuvekere | Udupi | Virajpet | Wadi | Yadgiri | Yelburga | Yellapur")
    s_a.append("Adimaly | Adoor | Agathy | Alappuzha | Alathur | Alleppey | Alwaye | Amini | Androth | Attingal | Badagara | Bitra | Calicut | Cannanore | Chetlet | Ernakulam | Idukki | Irinjalakuda | Kadamath | Kalpeni | Kalpetta | Kanhangad | Kanjirapally | Kannur | Karungapally | Kasargode | Kavarathy | Kiltan | Kochi | Koduvayur | Kollam | Kottayam | Kovalam | Kozhikode | Kunnamkulam | Malappuram | Mananthodi | Manjeri | Mannarghat | Mavelikkara | Minicoy | Munnar | Muvattupuzha | Nedumandad | Nedumgandam | Nilambur | Palai | Palakkad | Palghat | Pathaanamthitta | Pathanamthitta | Payyanur | Peermedu | Perinthalmanna | Perumbavoor | Punalur | Quilon | Ranni | Shertallai | Shoranur | Taliparamba | Tellicherry | Thiruvananthapuram | Thodupuzha | Thrissur | Tirur | Tiruvalla | Trichur | Trivandrum | Uppala | Vadakkanchery | Vikom | Wayanad")
    s_a.append("Agatti Island | Bingaram Island | Bitra Island | Chetlat Island | Kadmat Island | Kalpeni Island | Kavaratti Island | Kiltan Island | Lakshadweep Sea | Minicoy Island | North Island | South Island")
    s_a.append("Agar | Ajaigarh | Alirajpur | Amarpatan | Amarwada | Ambah | Anuppur | Arone | Ashoknagar | Ashta | Atner | Babaichichli | Badamalhera | Badarwsas | Badnagar | Badnawar | Badwani | Bagli | Baihar | Balaghat | Baldeogarh | Baldi | Bamori | Banda | Bandhavgarh | Bareli | Baroda | Barwaha | Barwani | Batkakhapa | Begamganj | Beohari | Berasia | Berchha | Betul | Bhainsdehi | Bhander | Bhanpura | Bhikangaon | Bhimpur | Bhind | Bhitarwar | Bhopal | Biaora | Bijadandi | Bijawar | Bijaypur | Bina | Birsa | Birsinghpur | Budhni | Burhanpur | Buxwaha | Chachaura | Chanderi | Chaurai | Chhapara | Chhatarpur | Chhindwara | Chicholi | Chitrangi | Churhat | Dabra | Damoh | Datia | Deori | Deosar | Depalpur | Dewas | Dhar | Dharampuri | Dindori | Gadarwara | Gairatganj | Ganjbasoda | Garoth | Ghansour | Ghatia | Ghatigaon | Ghorandogri | Ghughari | Gogaon | Gohad | Goharganj | Gopalganj | Gotegaon | Gourihar | Guna | Gunnore | Gwalior | Gyraspur | Hanumana | Harda | Harrai | Harsud | Hatta | Hoshangabad | Ichhawar | Indore | Isagarh | Itarsi | Jabalpur | Jabera | Jagdalpur | Jaisinghnagar | Jaithari | Jaitpur | Jaitwara | Jamai | Jaora | Jatara | Jawad | Jhabua | Jobat | Jora | Kakaiya | Kannod | Kannodi | Karanjia | Kareli | Karera | Karhal | Karpa | Kasrawad | Katangi | Katni | Keolari | Khachrod | Khajuraho | Khakner | Khalwa | Khandwa | Khaniadhana | Khargone | Khategaon | Khetia | Khilchipur | Khirkiya | Khurai | Kolaras | Kotma | Kukshi | Kundam | Kurwai | Kusmi | Laher | Lakhnadon | Lamta | Lanji | Lateri | Laundi | Maheshwar | Mahidpurcity | Maihar | Majhagwan | Majholi | Malhargarh | Manasa | Manawar | Mandla | Mandsaur | Manpur | Mauganj | Mawai | Mehgaon | Mhow | Morena | Multai | Mungaoli | Nagod | Nainpur | Narsingarh | Narsinghpur | Narwar | Nasrullaganj | Nateran | Neemuch | Niwari | Niwas | Nowgaon | Pachmarhi | Pandhana | Pandhurna | Panna | Parasia | Patan | Patera | Patharia | Pawai | Petlawad | Pichhore | Piparia | Pohari | Prabhapattan | Punasa | Pushprajgarh | Raghogarh | Raghunathpur | Rahatgarh | Raisen | Rajgarh | Rajpur | Ratlam | Rehli | Rewa | Sabalgarh | Sagar | Sailana | Sanwer | Sarangpur | Sardarpur | Satna | Saunsar | Sehore | Sendhwa | Seondha | Seoni | Seonimalwa | Shahdol | Shahnagar | Shahpur | Shajapur | Sheopur | Sheopurkalan | Shivpuri | Shujalpur | Sidhi | Sihora | Silwani | Singrauli | Sirmour | Sironj | Sitamau | Sohagpur | Sondhwa | Sonkatch | Susner | Tamia | Tarana | Tendukheda | Teonthar | Thandla | Tikamgarh | Timarani | Udaipura | Ujjain | Umaria | Umariapan | Vidisha | Vijayraghogarh | Waraseoni | Zhirnia")
    s_a.append("Achalpur | Aheri | Ahmednagar | Ahmedpur | Ajara | Akkalkot | Akola | Akole | Akot | Alibagh | Amagaon | Amalner | Ambad | Ambejogai | Amravati | Arjuni Merogaon | Arvi | Ashti | Atpadi | Aurangabad | Ausa | Babhulgaon | Balapur | Baramati | Barshi Takli | Barsi | Basmatnagar | Bassein | Beed | Bhadrawati | Bhamregadh | Bhandara | Bhir | Bhiwandi | Bhiwapur | Bhokar | Bhokardan | Bhoom | Bhor | Bhudargad | Bhusawal | Billoli | Brahmapuri | Buldhana | Butibori | Chalisgaon | Chamorshi | Chandgad | Chandrapur | Chandur | Chanwad | Chhikaldara | Chikhali | Chinchwad | Chiplun | Chopda | Chumur | Dahanu | Dapoli | Darwaha | Daryapur | Daund | Degloor | Delhi Tanda | Deogad | Deolgaonraja | Deori | Desaiganj | Dhadgaon | Dhanora | Dharani | Dhiwadi | Dhule | Dhulia | Digras | Dindori | Edalabad | Erandul | Etapalli | Gadhchiroli | Gadhinglaj | Gaganbavada | Gangakhed | Gangapur | Gevrai | Ghatanji | Golegaon | Gondia | Gondpipri | Goregaon | Guhagar | Hadgaon | Hatkangale | Hinganghat | Hingoli | Hingua | Igatpuri | Indapur | Islampur | Jalgaon | Jalna | Jamkhed | Jamner | Jath | Jawahar | Jintdor | Junnar | Kagal | Kaij | Kalamb | Kalamnuri | Kallam | Kalmeshwar | Kalwan | Kalyan | Kamptee | Kandhar | Kankavali | Kannad | Karad | Karjat | Karmala | Katol | Kavathemankal | Kedgaon | Khadakwasala | Khamgaon | Khed | Khopoli | Khultabad | Kinwat | Kolhapur | Kopargaon | Koregaon | Kudal | Kuhi | Kurkheda | Kusumba | Lakhandur | Langa | Latur | Lonar | Lonavala | Madangad | Madha | Mahabaleshwar | Mahad | Mahagaon | Mahasala | Mahaswad | Malegaon | Malgaon | Malgund | Malkapur | Malsuras | Malwan | Mancher | Mangalwedha | Mangaon | Mangrulpur | Manjalegaon | Manmad | Maregaon | Mehda | Mekhar | Mohadi | Mohol | Mokhada | Morshi | Mouda | Mukhed | Mul | Mumbai | Murbad | Murtizapur | Murud | Nagbhir | Nagpur | Nahavara | Nanded | Nandgaon | Nandnva | Nandurbar | Narkhed | Nashik | Navapur | Ner | Newasa | Nilanga | Niphad | Omerga | Osmanabad | Pachora | Paithan | Palghar | Pali | Pandharkawada | Pandharpur | Panhala | Paranda | Parbhani | Parner | Parola | Parseoni | Partur | Patan | Pathardi | Pathari | Patoda | Pauni | Peint | Pen | Phaltan | Pimpalner | Pirangut | Poladpur | Pune | Pusad | Pusegaon | Radhanagar | Rahuri | Raigad | Rajapur | Rajgurunagar | Rajura | Ralegaon | Ramtek | Ratnagiri | Raver | Risod | Roha | Sakarwadi | Sakoli | Sakri | Salekasa | Samudrapur | Sangamner | Sanganeshwar | Sangli | Sangola | Sanguem | Saoner | Saswad | Satana | Satara | Sawantwadi | Seloo | Shahada | Shahapur | Shahuwadi | Shevgaon | Shirala | Shirol | Shirpur | Shirur | Shirwal | Sholapur | Shri Rampur | Shrigonda | Shrivardhan | Sillod | Sinderwahi | Sindhudurg | Sindkheda | Sindkhedaraja | Sinnar | Sironcha | Soyegaon | Surgena | Talasari | Talegaon S.Ji Pant | Taloda | Tasgaon | Thane | Tirora | Tiwasa | Trimbak | Tuljapur | Tumsar | Udgir | Umarkhed | Umrane | Umrer | Urlikanchan | Vaduj | Velhe | Vengurla | Vijapur | Vita | Wada | Wai | Walchandnagar | Wani | Wardha | Warlydwarud | Warora | Washim | Wathar | Yavatmal | Yawal | Yeola | Yeotmal")
    s_a.append("Bishnupur | Chakpikarong | Chandel | Chattrik | Churachandpur | Imphal | Jiribam | Kakching | Kalapahar | Mao | Mulam | Parbung | Sadarhills | Saibom | Sempang | Senapati | Sochumer | Taloulong | Tamenglong | Thinghat | Thoubal | Ukhrul")
    s_a.append("Amlaren | Baghmara | Cherrapunjee | Dadengiri | Garo Hills | Jaintia Hills | Jowai | Khasi Hills | Khliehriat | Mariang | Mawkyrwat | Nongpoh | Nongstoin | Resubelpara | Ri Bhoi | Shillong | Tura | Williamnagar")
    s_a.append("Aizawl | Champhai | Demagiri | Kolasib | Lawngtlai | Lunglei | Mamit | Saiha | Serchhip")
    s_a.append("Dimapur | Jalukie | Kiphire | Kohima | Mokokchung | Mon | Phek | Tuensang | Wokha | Zunheboto")
    s_a.append("Anandapur | Angul | Anugul | Aska | Athgarh | Athmallik | Attabira | Bagdihi | Balangir | Balasore | Baleswar | Baliguda | Balugaon | Banaigarh | Bangiriposi | Barbil | Bargarh | Baripada | Barkot | Basta | Berhampur | Betanati | Bhadrak | Bhanjanagar | Bhawanipatna | Bhubaneswar | Birmaharajpur | Bisam Cuttack | Boriguma | Boudh | Buguda | Chandbali | Chhatrapur | Chhendipada | Cuttack | Daringbadi | Daspalla | Deodgarh | Deogarh | Dhanmandal | Dharamgarh | Dhenkanal | Digapahandi | Dunguripali | G. Udayagiri | Gajapati | Ganjam | Ghatgaon | Gudari | Gunupur | Hemgiri | Hindol | Jagatsinghapur | Jajpur | Jamankira | Jashipur | Jayapatna | Jeypur | Jharigan | Jharsuguda | Jujumura | Kalahandi | Kalimela | Kamakhyanagar | Kandhamal | Kantabhanji | Kantamal | Karanjia | Kashipur | Kendrapara | Kendujhar | Keonjhar | Khalikote | Khordha | Khurda | Komana | Koraput | Kotagarh | Kuchinda | Lahunipara | Laxmipur | M. Rampur | Malkangiri | Mathili | Mayurbhanj | Mohana | Motu | Nabarangapur | Naktideul | Nandapur | Narlaroad | Narsinghpur | Nayagarh | Nimapara | Nowparatan | Nowrangapur | Nuapada | Padampur | Paikamal | Palla Hara | Papadhandi | Parajang | Pardip | Parlakhemundi | Patnagarh | Pattamundai | Phiringia | Phulbani | Puri | Puruna Katak | R. Udayigiri | Rairakhol | Rairangpur | Rajgangpur | Rajkhariar | Rayagada | Rourkela | Sambalpur | Sohela | Sonapur | Soro | Subarnapur | Sunabeda | Sundergarh | Surada | T. Rampur | Talcher | Telkoi | Titlagarh | Tumudibandha | Udala | Umerkote")
    s_a.append("Bahur | Karaikal | Mahe | Pondicherry | Purnankuppam | Valudavur | Villianur | Yanam")
    s_a.append("Abohar | Ajnala | Amritsar | Balachaur | Barnala | Batala | Bathinda | Chandigarh | Dasua | Dinanagar | Faridkot | Fatehgarh Sahib | Fazilka | Ferozepur | Garhashanker | Goindwal | Gurdaspur | Guruharsahai | Hoshiarpur | Jagraon | Jalandhar | Jugial | Kapurthala | Kharar | Kotkapura | Ludhiana | Malaut | Malerkotla | Mansa | Moga | Muktasar | Nabha | Nakodar | Nangal | Nawanshahar | Nawanshahr | Pathankot | Patiala | Patti | Phagwara | Phillaur | Phulmandi | Quadian | Rajpura | Raman | Rayya | Ropar | Rupnagar | Samana | Samrala | Sangrur | Sardulgarh | Sarhind | SAS Nagar | Sultanpur Lodhi | Sunam | Tanda Urmar | Tarn Taran | Zira")
    s_a.append("Abu Road | Ahore | Ajmer | Aklera | Alwar | Amber | Amet | Anupgarh | Asind | Aspur | Atru | Bagidora | Bali | Bamanwas | Banera | Bansur | Banswara | Baran | Bari | Barisadri | Barmer | Baseri | Bassi | Baswa | Bayana | Beawar | Begun | Behror | Bhadra | Bharatpur | Bhilwara | Bhim | Bhinmal | Bikaner | Bilara | Bundi | Chhabra | Chhipaborad | Chirawa | Chittorgarh | Chohtan | Churu | Dantaramgarh | Dausa | Deedwana | Deeg | Degana | Deogarh | Deoli | Desuri | Dhariawad | Dholpur | Digod | Dudu | Dungarpur | Dungla | Fatehpur | Gangapur | Gangdhar | Gerhi | Ghatol | Girwa | Gogunda | Hanumangarh | Hindaun | Hindoli | Hurda | Jahazpur | Jaipur | Jaisalmer | Jalore | Jhalawar | Jhunjhunu | Jodhpur | Kaman | Kapasan | Karauli | Kekri | Keshoraipatan | Khandar | Kherwara | Khetri | Kishanganj | Kishangarh | Kishangarhbas | Kolayat | Kota | Kotputli | Kotra | Kotri | Kumbalgarh | Kushalgarh | Ladnun | Ladpura | Lalsot | Laxmangarh | Lunkaransar | Mahuwa | Malpura | Malvi | Mandal | Mandalgarh | Mandawar | Mangrol | Marwar-Jn | Merta | Nadbai | Nagaur | Nainwa | Nasirabad | Nathdwara | Nawa | Neem Ka Thana | Newai | Nimbahera | Nohar | Nokha | Onli | Osian | Pachpadara | Pachpahar | Padampur | Pali | Parbatsar | Phagi | Phalodi | Pilani | Pindwara | Pipalda | Pirawa | Pokaran | Pratapgarh | Raipur | Raisinghnagar | Rajgarh | Rajsamand | Ramganj Mandi | Ramgarh | Rashmi | Ratangarh | Reodar | Rupbas | Sadulshahar | Sagwara | Sahabad | Salumber | Sanchore | Sangaria | Sangod | Sapotra | Sarada | Sardarshahar | Sarwar | Sawai Madhopur | Shahapura | Sheo | Sheoganj | Shergarh | Sikar | Sirohi | Siwana | Sojat | Sri Dungargarh | Sri Ganganagar | Sri Karanpur | Sri Madhopur | Sujangarh | Taranagar | Thanaghazi | Tibbi | Tijara | Todaraisingh | Tonk | Udaipur | Udaipurwati | Uniayara | Vallabhnagar | Viratnagar")
    s_a.append("Barmiak | Be | Bhurtuk | Chhubakha | Chidam | Chubha | Chumikteng | Dentam | Dikchu | Dzongri | Gangtok | Gauzing | Gyalshing | Hema | Kerung | Lachen | Lachung | Lema | Lingtam | Lungthu | Mangan | Namchi | Namthang | Nanga | Nantang | Naya Bazar | Padamachen | Pakhyong | Pemayangtse | Phensang | Rangli | Rinchingpong | Sakyong | Samdong | Singtam | Siniolchu | Sombari | Soreng | Sosing | Tekhug | Temi | Tsetang | Tsomgo | Tumlong | Yangang | Yumtang")
    s_a.append("Ambasamudram | Anamali | Arakandanallur | Arantangi | Aravakurichi | Ariyalur | Arkonam | Arni | Aruppukottai | Attur | Avanashi | Batlagundu | Bhavani | Chengalpattu | Chengam | Chennai | Chidambaram | Chingleput | Coimbatore | Courtallam | Cuddalore | Cumbum | Denkanikoitah | Devakottai | Dharampuram | Dharmapuri | Dindigul | Erode | Gingee | Gobichettipalayam | Gudalur | Gudiyatham | Harur | Hosur | Jayamkondan | Kallkurichi | Kanchipuram | Kangayam | Kanyakumari | Karaikal | Karaikudi | Karur | Keeranur | Kodaikanal | Kodumudi | Kotagiri | Kovilpatti | Krishnagiri | Kulithalai | Kumbakonam | Kuzhithurai | Madurai | Madurantgam | Manamadurai | Manaparai | Mannargudi | Mayiladuthurai | Mayiladutjurai | Mettupalayam | Metturdam | Mudukulathur | Mulanur | Musiri | Nagapattinam | Nagarcoil | Namakkal | Nanguneri | Natham | Neyveli | Nilgiris | Oddanchatram | Omalpur | Ootacamund | Ooty | Orathanad | Palacode | Palani | Palladum | Papanasam | Paramakudi | Pattukottai | Perambalur | Perundurai | Pollachi | Polur | Pondicherry | Ponnamaravathi | Ponneri | Pudukkottai | Rajapalayam | Ramanathapuram | Rameshwaram | Ranipet | Rasipuram | Salem | Sankagiri | Sankaran | Sathiyamangalam | Sivaganga | Sivakasi | Sriperumpudur | Srivaikundam | Tenkasi | Thanjavur | Theni | Thirumanglam | Thiruraipoondi | Thoothukudi | Thuraiyure | Tindivanam | Tiruchendur | Tiruchengode | Tiruchirappalli | Tirunelvelli | Tirupathur | Tirupur | Tiruttani | Tiruvallur | Tiruvannamalai | Tiruvarur | Tiruvellore | Tiruvettipuram | Trichy | Tuticorin | Udumalpet | Ulundurpet | Usiliampatti | Uthangarai | Valapady | Valliyoor | Vaniyambadi | Vedasandur | Vellore | Velur | Vilathikulam | Villupuram | Virudhachalam | Virudhunagar | Wandiwash | Yercaud")
    s_a.append("Agartala | Ambasa | Bampurbari | Belonia | Dhalai | Dharam Nagar | Kailashahar | Kamal Krishnabari | Khopaiyapara | Khowai | Phuldungsei | Radha Kishore Pur | Tripura")
    s_a.append("Achhnera | Agra | Akbarpur | Aliganj | Aligarh | Allahabad | Ambedkar Nagar | Amethi | Amiliya | Amroha | Anola | Atrauli | Auraiya | Azamgarh | Baberu | Badaun | Baghpat | Bagpat | Baheri | Bahraich | Ballia | Balrampur | Banda | Bansdeeh | Bansgaon | Bansi | Barabanki | Bareilly | Basti | Bhadohi | Bharthana | Bharwari | Bhogaon | Bhognipur | Bidhuna | Bijnore | Bikapur | Bilari | Bilgram | Bilhaur | Bindki | Bisalpur | Bisauli | Biswan | Budaun | Budhana | Bulandshahar | Bulandshahr | Capianganj | Chakia | Chandauli | Charkhari | Chhata | Chhibramau | Chirgaon | Chitrakoot | Chunur | Dadri | Dalmau | Dataganj | Debai | Deoband | Deoria | Derapur | Dhampur | Domariyaganj | Dudhi | Etah | Etawah | Faizabad | Farrukhabad | Fatehpur | Firozabad | Garauth | Garhmukteshwar | Gautam Buddha Nagar | Ghatampur | Ghaziabad | Ghazipur | Ghosi | Gonda | Gorakhpur | Gunnaur | Haidergarh | Hamirpur | Hapur | Hardoi | Harraiya | Hasanganj | Hasanpur | Hathras | Jalalabad | Jalaun | Jalesar | Jansath | Jarar | Jasrana | Jaunpur | Jhansi | Jyotiba Phule Nagar | Kadipur | Kaimganj | Kairana | Kaisarganj | Kalpi | Kannauj | Kanpur | Karchhana | Karhal | Karvi | Kasganj | Kaushambi | Kerakat | Khaga | Khair | Khalilabad | Kheri | Konch | Kumaon | Kunda | Kushinagar | Lalganj | Lalitpur | Lucknow | Machlishahar | Maharajganj | Mahoba | Mainpuri | Malihabad | Mariyahu | Math | Mathura | Mau | Maudaha | Maunathbhanjan | Mauranipur | Mawana | Meerut | Mehraun | Meja | Mirzapur | Misrikh | Modinagar | Mohamdabad | Mohamdi | Moradabad | Musafirkhana | Muzaffarnagar | Nagina | Najibabad | Nakur | Nanpara | Naraini | Naugarh | Nawabganj | Nighasan | Noida | Orai | Padrauna | Pahasu | Patti | Pharenda | Phoolpur | Phulpur | Pilibhit | Pitamberpur | Powayan | Pratapgarh | Puranpur | Purwa | Raibareli | Rampur | Ramsanehi Ghat | Rasara | Rath | Robertsganj | Sadabad | Safipur | Sagri | Saharanpur | Sahaswan | Sahjahanpur | Saidpur | Salempur | Salon | Sambhal | Sandila | Sant Kabir Nagar | Sant Ravidas Nagar | Sardhana | Shahabad | Shahganj | Shahjahanpur | Shikohabad | Shravasti | Siddharthnagar | Sidhauli | Sikandra Rao | Sikandrabad | Sitapur | Siyana | Sonbhadra | Soraon | Sultanpur | Tanda | Tarabganj | Tilhar | Unnao | Utraula | Varanasi | Zamania")
    s_a.append("Almora | Bageshwar | Bhatwari | Chakrata | Chamoli | Champawat | Dehradun | Deoprayag | Dharchula | Dunda | Haldwani | Haridwar | Joshimath | Karan Prayag | Kashipur | Khatima | Kichha | Lansdown | Munsiari | Mussoorie | Nainital | Pantnagar | Partapnagar | Pauri Garhwal | Pithoragarh | Purola | Rajgarh | Ranikhet | Roorkee | Rudraprayag | Tehri Garhwal | Udham Singh Nagar | Ukhimath | Uttarkashi")
    s_a.append("Adra | Alipurduar | Amlagora | Arambagh | Asansol | Balurghat | Bankura | Bardhaman | Basirhat | Berhampur | Bethuadahari | Birbhum | Birpara | Bishanpur | Bolpur | Bongoan | Bulbulchandi | Burdwan | Calcutta | Canning | Champadanga | Contai | Cooch Behar | Daimond Harbour | Dalkhola | Dantan | Darjeeling | Dhaniakhali | Dhuliyan | Dinajpur | Dinhata | Durgapur | Gangajalghati | Gangarampur | Ghatal | Guskara | Habra | Haldia | Harirampur | Harishchandrapur | Hooghly | Howrah | Islampur | Jagatballavpur | Jalpaiguri | Jhalda | Jhargram | Kakdwip | Kalchini | Kalimpong | Kalna | Kandi | Karimpur | Katwa | Kharagpur | Khatra | Krishnanagar | Mal Bazar | Malda | Manbazar | Mathabhanga | Medinipur | Mekhliganj | Mirzapur | Murshidabad | Nadia | Nagarakata | Nalhati | Nayagarh | Parganas | Purulia | Raiganj | Rampur Hat | Ranaghat | Seharabazar | Siliguri | Suri | Takipur | Tamluk")

    for state in range(0,len(state_arr)):
        stte = State(state=state_arr[state])
        stte.save()
        temp = s_a[state].split(' | ')
        for city in temp:        
            ct = City(city=city,state=stte)
            ct.save()